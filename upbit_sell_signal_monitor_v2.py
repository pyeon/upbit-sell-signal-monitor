#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
업비트 실시간 매도 신호 모니터링 시스템 v2.0
개선 사항:
- 10분봉 + 60분봉 혼합 분석으로 급격한 변동 감지
- Config에서 모든 타임프레임과 임계값 조정 가능
- 30분 주기로 실행 가능 (급락을 더 빨리 감지)
"""

import pyupbit
import pandas as pd
import numpy as np
import requests
import time
from datetime import datetime, timedelta
import pytz
import ta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import warnings
warnings.filterwarnings('ignore')

# 한국 시간대 설정
KST = pytz.timezone('Asia/Seoul')

# 설정 파일 불러오기
try:
    from config import *
except ImportError:
    print("❌ config.py 파일이 없습니다!")
    print("📝 config.example.py를 config.py로 복사하고 설정을 입력하세요.")
    exit(1)

# ============================================
# 시간 관련 함수
# ============================================

def get_kst_now():
    """한국 시간 반환"""
    return datetime.now(KST)

def format_kst_time(dt=None):
    """한국 시간 포맷팅"""
    if dt is None:
        dt = get_kst_now()
    return dt.strftime('%Y-%m-%d %H:%M:%S')

# ============================================
# 텔레그램 전송 함수
# ============================================

def send_telegram(message, parse_mode=None):
    """텔레그램 메시지 전송"""
    try:
        url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
        data = {
            "chat_id": CHAT_ID,
            "text": message
        }
        if parse_mode:
            data["parse_mode"] = parse_mode
            
        response = requests.post(url, data=data, timeout=10)
        return response.json()
    except Exception as e:
        print(f"텔레그램 전송 실패: {e}")
        return None

# ============================================
# 급등 후 하락 패턴 분석 (개선)
# ============================================

def analyze_price_pattern(coin):
    """
    급등 후 급락 패턴 감지 (개선된 버전)
    - 단기(10분봉): 급격한 변동 감지
    - 중기(60분봉): 전체 흐름 파악
    - 장기(일봉): 추세 확인
    """
    try:
        # 10분봉 데이터 (최근 12시간 = 72개)
        df_10m = pyupbit.get_ohlcv(coin, interval="minute10", count=MINUTE_10_COUNT)
        if df_10m is None or len(df_10m) < 30:
            return None
        
        # 60분봉 데이터 (최근 24시간)
        df_60m = pyupbit.get_ohlcv(coin, interval="minute60", count=MINUTE_60_COUNT)
        if df_60m is None or len(df_60m) < 12:
            return None
        
        # 일봉 데이터 (최근 30일)
        df_day = pyupbit.get_ohlcv(coin, interval="day", count=30)
        if df_day is None or len(df_day) < 20:
            return None
        
        current_price = df_10m['close'].iloc[-1]
        
        # ===== 1. 단기 급락 감지 (10분봉) =====
        # 최근 N개 봉 중 최고가 (Config에서 조정 가능)
        recent_candles = df_10m.tail(QUICK_DROP_LOOKBACK)
        recent_high = recent_candles['high'].max()
        quick_drop = ((recent_high - current_price) / recent_high) * 100
        
        # 급락 발생 시점 (몇 개 봉 전?)
        high_idx = recent_candles['high'].idxmax()
        candles_since_high = len(recent_candles) - recent_candles.index.get_loc(high_idx) - 1
        minutes_since_high = candles_since_high * 10
        
        # ===== 2. 중기 추세 (60분봉) =====
        # 최근 12시간 최고가
        high_12h = df_60m['high'].max()
        drop_from_high_12h = ((high_12h - current_price) / high_12h) * 100
        
        # 최근 6시간 상승률
        if len(df_60m) >= 7:
            price_6h_ago = df_60m['close'].iloc[-7]
            surge_6h = ((current_price - price_6h_ago) / price_6h_ago) * 100
        else:
            surge_6h = 0
        
        # 최근 1시간 변화율
        price_1h_ago = df_60m['close'].iloc[-2]
        change_1h = ((current_price - price_1h_ago) / price_1h_ago) * 100
        
        # ===== 3. 장기 추세 (일봉) =====
        price_7d_ago = df_day['close'].iloc[-8] if len(df_day) >= 8 else df_day['close'].iloc[0]
        change_7d = ((current_price - price_7d_ago) / price_7d_ago) * 100
        
        # ===== 4. 변동성 체크 (10분봉 기준) =====
        # 최근 N개 봉의 평균 변동률
        recent_volatility = []
        for i in range(1, min(VOLATILITY_CHECK_CANDLES + 1, len(df_10m))):
            change = abs((df_10m['close'].iloc[-i] - df_10m['close'].iloc[-(i+1)]) / df_10m['close'].iloc[-(i+1)]) * 100
            recent_volatility.append(change)
        
        avg_volatility = np.mean(recent_volatility) if recent_volatility else 0
        
        return {
            'current_price': current_price,
            
            # 단기 (10분봉)
            'quick_drop': quick_drop,  # 최근 N개 봉 중 최고가 대비 하락
            'minutes_since_high': minutes_since_high,  # 고점 이후 경과 시간
            'recent_high': recent_high,
            
            # 중기 (60분봉)
            'high_12h': high_12h,
            'drop_from_high_12h': drop_from_high_12h,
            'surge_6h': surge_6h,
            'change_1h': change_1h,
            
            # 장기 (일봉)
            'change_7d': change_7d,
            
            # 변동성
            'avg_volatility': avg_volatility
        }
    except Exception as e:
        print(f"가격 패턴 분석 오류: {e}")
        return None

# ============================================
# 거래량 분석 (하락 전환)
# ============================================

def analyze_volume_decline(coin):
    """거래량 감소 및 다이버전스 분석"""
    try:
        # 일봉 데이터
        df = pyupbit.get_ohlcv(coin, interval="day", count=30)
        if df is None or len(df) < 20:
            return None
        
        current_volume = df['volume'].iloc[-1]
        volume_ma_20 = df['volume'].rolling(20).mean().iloc[-1]
        
        # 1. 거래량 MA 대비
        volume_ratio = current_volume / volume_ma_20
        
        # 2. 거래량 감소 추세 (최근 N일)
        volume_trend = []
        for i in range(1, VOLUME_DECLINE_DAYS + 1):
            if len(df) > i:
                volume_trend.append(df['volume'].iloc[-i])
        
        volume_declining = all(volume_trend[i] < volume_trend[i-1] for i in range(1, len(volume_trend))) if len(volume_trend) > 1 else False
        
        # 3. 가격-거래량 다이버전스
        lookback = DIVERGENCE_LOOKBACK_DAYS
        if len(df) > lookback:
            price_change = ((df['close'].iloc[-1] - df['close'].iloc[-(lookback+1)]) / df['close'].iloc[-(lookback+1)]) * 100
            volume_change = ((current_volume - df['volume'].iloc[-(lookback+1)]) / df['volume'].iloc[-(lookback+1)]) * 100
        else:
            price_change = 0
            volume_change = 0
        
        # 가격은 상승했지만 거래량이 감소 = 약세 다이버전스
        divergence_signal = (price_change > DIVERGENCE_PRICE_THRESHOLD) and (volume_change < DIVERGENCE_VOLUME_THRESHOLD)
        
        return {
            'volume_ratio': volume_ratio,
            'volume_declining': volume_declining,
            'divergence_signal': divergence_signal,
            'price_change': price_change,
            'volume_change': volume_change
        }
    except Exception as e:
        print(f"거래량 분석 오류: {e}")
        return None

# ============================================
# 호가창 분석 (매도 우세)
# ============================================

def analyze_orderbook_sell(coin):
    """호가창 매도 압력 분석"""
    try:
        orderbook = pyupbit.get_orderbook(coin)
        if orderbook is None or not isinstance(orderbook, list) or len(orderbook) == 0:
            return None
        
        orderbook_data = orderbook[0]
        
        if 'orderbook_units' not in orderbook_data:
            return None
        
        units = orderbook_data['orderbook_units']
        
        # 매수/매도 총 물량
        total_bid_size = sum([item.get('bid_size', 0) for item in units])
        total_ask_size = sum([item.get('ask_size', 0) for item in units])
        
        # 매도/매수 비율 (매도가 크면 높음)
        ask_bid_ratio = total_ask_size / total_bid_size if total_bid_size > 0 else 0
        
        # 최상단 매도 물량
        top_bid = units[0].get('bid_size', 0) if len(units) > 0 else 0
        top_ask = units[0].get('ask_size', 0) if len(units) > 0 else 0
        
        return {
            'total_bid': total_bid_size,
            'total_ask': total_ask_size,
            'ask_bid_ratio': ask_bid_ratio,
            'top_bid': top_bid,
            'top_ask': top_ask
        }
    except Exception as e:
        return None

# ============================================
# 기술적 지표 (매도 신호)
# ============================================

def calculate_sell_indicators(coin):
    """매도 관련 기술적 지표"""
    try:
        df = pyupbit.get_ohlcv(coin, interval="day", count=100)
        if df is None or len(df) < 50:
            return None
        
        # 1. RSI (과매수)
        rsi = ta.momentum.RSIIndicator(df['close'], window=14).rsi().iloc[-1]
        rsi_signal = "과매수" if rsi > RSI_OVERBOUGHT else "고점권" if rsi > RSI_HIGH else "중립"
        
        # 2. MACD (데드크로스)
        macd = ta.trend.MACD(df['close'])
        macd_line = macd.macd().iloc[-1]
        signal_line = macd.macd_signal().iloc[-1]
        macd_hist = macd.macd_diff().iloc[-1]
        macd_signal = "데드크로스" if macd_line < signal_line and macd_hist < 0 else "약세전환" if macd_line < signal_line else "중립"
        
        # 3. 볼린저 밴드 (상단 이탈)
        bollinger = ta.volatility.BollingerBands(df['close'])
        bb_high = bollinger.bollinger_hband().iloc[-1]
        bb_low = bollinger.bollinger_lband().iloc[-1]
        current_price = df['close'].iloc[-1]
        
        # 상단 터치 후 하락 확인
        price_pct = (current_price - bb_low) / (bb_high - bb_low) * 100
        
        if current_price >= bb_high:
            bb_signal = "상단이탈"
        elif price_pct > BB_HIGH_THRESHOLD:
            bb_signal = "상단근접"
        else:
            bb_signal = "중립"
        
        # 4. 이동평균선 (하향 전환)
        ma5 = df['close'].rolling(5).mean().iloc[-1]
        ma20 = df['close'].rolling(20).mean().iloc[-1]
        ma_signal = "하향돌파" if ma5 < ma20 else "하향접근" if current_price < ma5 else "중립"
        
        # 5. 스토캐스틱 (과매수)
        stoch = ta.momentum.StochasticOscillator(df['high'], df['low'], df['close'])
        stoch_k = stoch.stoch().iloc[-1]
        stoch_signal = "과매수" if stoch_k > STOCH_OVERBOUGHT else "고점권" if stoch_k > STOCH_HIGH else "중립"
        
        return {
            'rsi': rsi,
            'rsi_signal': rsi_signal,
            'macd_signal': macd_signal,
            'bb_signal': bb_signal,
            'bb_position': price_pct,
            'ma_signal': ma_signal,
            'stoch': stoch_k,
            'stoch_signal': stoch_signal,
            'current_price': current_price
        }
    except Exception as e:
        return None

# ============================================
# 매도 신호 강도 계산 (개선)
# ============================================

def calculate_sell_signal_strength(pattern_data, volume_data, orderbook_data, indicators):
    """10개 지표 기반 매도 신호 강도 계산 (급락 감지 추가)"""
    score = 0
    signals = []
    
    # 가격 패턴 분석 (4개 - 추가됨!)
    if pattern_data:
        # 1. 단기 급락 (10분봉 기준 - NEW!)
        if pattern_data['quick_drop'] > QUICK_DROP_THRESHOLD:
            score += 1
            signals.append(f"✅ 단기급락 {pattern_data['minutes_since_high']}분전 -{pattern_data['quick_drop']:.1f}%")
        
        # 2. 중기 고점 대비 하락 (12시간)
        if pattern_data['drop_from_high_12h'] > DROP_FROM_HIGH_12H_THRESHOLD:
            score += 1
            signals.append(f"✅ 12시간 고점대비 -{pattern_data['drop_from_high_12h']:.1f}%")
        
        # 3. 급등 후 하락 전환
        if pattern_data['surge_6h'] > SURGE_6H_THRESHOLD and pattern_data['change_1h'] < CHANGE_1H_THRESHOLD:
            score += 1
            signals.append("✅ 급등 후 하락전환")
        
        # 4. 변동성 급증 (급락 직전 신호)
        if pattern_data['avg_volatility'] > VOLATILITY_THRESHOLD:
            score += 1
            signals.append(f"✅ 고변동성 {pattern_data['avg_volatility']:.1f}%")
    
    # 거래량 분석 (2개)
    if volume_data:
        # 5. 거래량 감소 추세
        if volume_data['volume_declining']:
            score += 1
            signals.append("✅ 거래량 감소 추세")
        
        # 6. 가격-거래량 다이버전스
        if volume_data['divergence_signal']:
            score += 1
            signals.append("✅ 약세 다이버전스")
    
    # 호가창 (1개)
    if orderbook_data:
        # 7. 매도벽 우세
        if orderbook_data['ask_bid_ratio'] > ORDERBOOK_THRESHOLD:
            score += 1
            signals.append("✅ 매도벽 우세")
    
    # 기술적 지표 (3개)
    if indicators:
        # 8. RSI 과매수
        if indicators['rsi'] > RSI_OVERBOUGHT:
            score += 1
            signals.append("✅ RSI 과매수")
        
        # 9. MACD 데드크로스
        if indicators['macd_signal'] == "데드크로스":
            score += 1
            signals.append("✅ MACD 데드크로스")
        
        # 10. 볼린저 상단 이탈
        if indicators['bb_signal'] in ["상단이탈", "상단근접"]:
            score += 1
            signals.append("✅ 볼린저 상단권")
    
    return score, signals

# ============================================
# 매도 단계 판단
# ============================================

def determine_sell_stage(score):
    """매도 단계 3단계 구분"""
    if score >= SELL_STAGE_IMMEDIATE:
        return {
            'stage': '즉시매도',
            'emoji': '🔴',
            'stars': '⭐' * 5,
            'color': 'red',
            'action': '즉시 매도 권장'
        }
    elif score >= SELL_STAGE_PREPARE:
        return {
            'stage': '매도준비',
            'emoji': '🟠',
            'stars': '⭐' * 3,
            'color': 'orange',
            'action': '일부 매도 고려'
        }
    elif score >= SELL_STAGE_REVIEW:
        return {
            'stage': '매도검토',
            'emoji': '🟡',
            'stars': '⭐' * 2,
            'color': 'yellow',
            'action': '주의 관찰 필요'
        }
    else:
        return None

# ============================================
# 텔레그램 메시지 포맷팅 (매도용)
# ============================================

def format_sell_telegram_message(coin, score, signals, pattern_data, volume_data, orderbook_data, indicators):
    """텔레그램 매도 메시지 생성"""
    
    stage_info = determine_sell_stage(score)
    if stage_info is None:
        return None
    
    coin_name = coin.replace("KRW-", "")
    
    # 메시지 구성
    message = f"{stage_info['emoji']} [{coin_name}] {stage_info['stage']} {stage_info['stars']}\n"
    message += "━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"💰 현재가: {pattern_data['current_price']:,.0f}원\n"
    message += f"🎯 권장행동: {stage_info['action']}\n\n"
    
    message += "【 가격 패턴 분석 】\n"
    
    if pattern_data:
        # 단기 급락 강조 (NEW!)
        if pattern_data['quick_drop'] > 3:
            message += f"🚨 단기급락: {pattern_data['minutes_since_high']}분 전 고점({pattern_data['recent_high']:,.0f}원) 대비 -{pattern_data['quick_drop']:.1f}%\n"
        
        if pattern_data['drop_from_high_12h'] > 5:
            message += f"📉 12시간 고점({pattern_data['high_12h']:,.0f}원) 대비: -{pattern_data['drop_from_high_12h']:.1f}%\n"
        
        if abs(pattern_data['surge_6h']) > 10:
            surge_emoji = "📈" if pattern_data['surge_6h'] > 0 else "📉"
            message += f"{surge_emoji} 6시간 변화: {pattern_data['surge_6h']:+.1f}%\n"
        
        if abs(pattern_data['change_1h']) > 2:
            change_emoji = "⚠️" if pattern_data['change_1h'] < 0 else "📊"
            message += f"{change_emoji} 1시간 변화: {pattern_data['change_1h']:+.1f}%\n"
        
        if pattern_data['avg_volatility'] > 2:
            message += f"⚡ 변동성: {pattern_data['avg_volatility']:.1f}% (최근 {VOLATILITY_CHECK_CANDLES}개 봉)\n"
    
    message += "\n【 거래량 분석 】\n"
    
    if volume_data:
        if volume_data['volume_declining']:
            message += f"⚠️ 거래량: {VOLUME_DECLINE_DAYS}일 연속 감소 ▶ 상승동력 약화\n"
        
        if volume_data['divergence_signal']:
            message += f"⚡ 약세 다이버전스:\n"
            message += f"   └ 가격 {volume_data['price_change']:+.1f}%, 거래량 {volume_data['volume_change']:+.1f}%\n"
            message += f"   └ 가격상승에도 거래량 감소 ▶ 매도 신호\n"
        
        message += f"📊 거래량 MA 대비: {volume_data['volume_ratio']:.2f}배\n"
    
    if orderbook_data:
        message += f"\n📊 호가창: 매도/매수 비율 {orderbook_data['ask_bid_ratio']:.2f}\n"
        if orderbook_data['ask_bid_ratio'] > ORDERBOOK_THRESHOLD:
            message += f"   └ 매도벽 우세 ▶ 하방 압력\n"
    
    message += "\n【 기술적 지표 】\n"
    
    if indicators:
        rsi_emoji = "✅" if indicators['rsi'] > RSI_OVERBOUGHT else "⚠️" if indicators['rsi'] > RSI_HIGH else "📊"
        message += f"{rsi_emoji} RSI: {indicators['rsi']:.1f} → {indicators['rsi_signal']}\n"
        
        macd_emoji = "✅" if indicators['macd_signal'] == '데드크로스' else "📊"
        message += f"{macd_emoji} MACD: {indicators['macd_signal']}\n"
        
        bb_emoji = "✅" if indicators['bb_signal'] in ['상단이탈', '상단근접'] else "📊"
        message += f"{bb_emoji} 볼린저: {indicators['bb_signal']} ({indicators['bb_position']:.0f}%)\n"
        
        ma_emoji = "✅" if indicators['ma_signal'] == '하향돌파' else "📊"
        message += f"{ma_emoji} 이동평균: {indicators['ma_signal']}\n"
        
        stoch_emoji = "✅" if indicators['stoch'] > STOCH_OVERBOUGHT else "⚠️" if indicators['stoch'] > STOCH_HIGH else "📊"
        message += f"{stoch_emoji} 스토캐스틱: {indicators['stoch']:.1f} → {indicators['stoch_signal']}\n"
    
    message += "\n━━━━━━━━━━━━━━━━━━━━━\n"
    message += f"🎯 종합판단: {score}/10 지표 일치\n"
    message += f"⏰ 발생시각(KST): {format_kst_time()}"
    
    return message

# ============================================
# 엑셀 저장 함수
# ============================================

def save_to_excel(coin, score, stage, pattern_data, volume_data, orderbook_data, indicators):
    """엑셀에 매도 신호 저장"""
    try:
        filename = "upbit_sell_signals_v2.xlsx"
        
        # 기존 파일 열기 또는 새로 생성
        try:
            wb = load_workbook(filename)
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active
            ws.title = "매도 신호"
            
            # 헤더 작성
            headers = ['시간(KST)', '코인', '매도단계', '신호강도', '현재가', '단기급락', 
                      '12시간고점대비', '6시간변화', '거래량추세', '다이버전스', '호가비율', 
                      'RSI', 'MACD', '볼린저', 'MA', '스토캐스틱']
            ws.append(headers)
            
            # 헤더 스타일
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        # 데이터 추가
        row_data = [
            format_kst_time(),
            coin.replace('KRW-', ''),
            stage,
            f"{score}/10",
            pattern_data['current_price'] if pattern_data else '',
            f"-{pattern_data['quick_drop']:.1f}% ({pattern_data['minutes_since_high']}분)" if pattern_data else '',
            f"-{pattern_data['drop_from_high_12h']:.1f}%" if pattern_data else '',
            f"{pattern_data['surge_6h']:+.1f}%" if pattern_data else '',
            "감소" if volume_data and volume_data['volume_declining'] else "정상",
            "있음" if volume_data and volume_data['divergence_signal'] else "없음",
            f"{orderbook_data['ask_bid_ratio']:.2f}" if orderbook_data else '',
            f"{indicators['rsi']:.1f}" if indicators else '',
            indicators['macd_signal'] if indicators else '',
            indicators['bb_signal'] if indicators else '',
            indicators['ma_signal'] if indicators else '',
            f"{indicators['stoch']:.1f}" if indicators else ''
        ]
        
        ws.append(row_data)
        
        # 100개 행만 유지
        if ws.max_row > 101:
            ws.delete_rows(2, ws.max_row - 101)
        
        wb.save(filename)
        print(f"✅ 엑셀 저장 완료: {coin}")
        
    except Exception as e:
        print(f"엑셀 저장 오류: {e}")

# ============================================
# 메인 스캔 함수
# ============================================

def scan_sell_signals():
    """매도 신호 스캔"""
    print(f"\n{'='*50}")
    print(f"🔍 매도 신호 스캔 시작 (v2.0): {format_kst_time()}")
    print(f"{'='*50}\n")
    
    # 원화 마켓 코인 리스트
    tickers = pyupbit.get_tickers(fiat="KRW")
    print(f"📊 총 {len(tickers)}개 코인 분석 중...\n")
    
    signal_count = 0
    
    for idx, coin in enumerate(tickers, 1):
        try:
            # 진행률 표시
            if idx % 50 == 0:
                print(f"진행률: {idx}/{len(tickers)} ({idx/len(tickers)*100:.1f}%)")
            
            # 1단계: 가격 패턴 분석 (개선된 버전)
            pattern_data = analyze_price_pattern(coin)
            if not pattern_data:
                continue
            
            # 필터링: 최소한의 변동이 있는 코인만
            if pattern_data['quick_drop'] < MIN_QUICK_DROP and pattern_data['drop_from_high_12h'] < MIN_DROP_12H:
                continue
            
            print(f"🔎 {coin}: 가격 변동 감지 - 정밀 분석 중...")
            
            # 2단계: 거래량 분석
            volume_data = analyze_volume_decline(coin)
            
            # 3단계: 호가창 분석
            orderbook_data = analyze_orderbook_sell(coin)
            
            # 4단계: 기술적 지표
            indicators = calculate_sell_indicators(coin)
            
            # 5단계: 신호 강도 계산
            score, signals = calculate_sell_signal_strength(
                pattern_data, volume_data, orderbook_data, indicators
            )
            
            # 6단계: 매도 신호 발송
            if score >= SELL_STAGE_REVIEW:
                signal_count += 1
                stage_info = determine_sell_stage(score)
                
                if stage_info:
                    # 텔레그램 메시지
                    message = format_sell_telegram_message(
                        coin, score, signals, pattern_data, volume_data, 
                        orderbook_data, indicators
                    )
                    if message:
                        send_telegram(message)
                        print(f"✅ 매도신호 발송: {coin} ({stage_info['stage']}, {score}/10)")
                    
                    # 엑셀 저장
                    save_to_excel(
                        coin, score, stage_info['stage'], pattern_data, 
                        volume_data, orderbook_data, indicators
                    )
            
            # API 제한 방지
            time.sleep(0.1)
            
        except Exception as e:
            print(f"❌ {coin} 분석 오류: {e}")
            continue
    
    print(f"\n{'='*50}")
    print(f"✅ 스캔 완료: 총 {signal_count}개 매도신호 발견")
    print(f"{'='*50}\n")

# ============================================
# 메인 실행
# ============================================

def main():
    """메인 실행 함수"""
    print("""
    ╔══════════════════════════════════════╗
    ║   업비트 매도 신호 모니터링 v2.0     ║
    ║      (10분봉 급락 감지 추가)         ║
    ║          (한국시간 기준)             ║
    ╚══════════════════════════════════════╝
    """)
    
    # 텔레그램 연결 테스트
    print(f"📱 텔레그램 연결 테스트 중... (Chat ID: {CHAT_ID})")
    test_result = send_telegram(f"🔴 업비트 매도 신호 모니터링 v2.0 시작! (KST: {format_kst_time()})")
    
    if test_result and test_result.get('ok'):
        print("✅ 텔레그램 연결 성공!\n")
    else:
        print("❌ 텔레그램 연결 실패!")
        print(f"응답: {test_result}\n")
        print("⚠️  그래도 스캔을 진행합니다...\n")
    
    # 메인 스캔 실행
    try:
        scan_sell_signals()
        
    except KeyboardInterrupt:
        print("\n\n🛑 매도 모니터링 중지됨")
        send_telegram(f"🛑 업비트 매도 신호 모니터링 종료 (KST: {format_kst_time()})")

if __name__ == "__main__":
    main()
